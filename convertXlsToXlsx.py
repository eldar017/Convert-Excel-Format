
import uuid
import xlrd
from xlrd.sheet import ctype_text
from openpyxl import Workbook

wbb = Workbook()
ws = wbb.active
try:
    #xl_workbook = xlrd.open_workbook('c:\om\sample-xls-file-for-testing.xls')
    xl_workbook = xlrd.open_workbook('c:\om\Transactions_14_04_2020.xls')
    sheet_names = xl_workbook.sheet_names()
    print(sheet_names)

    xl_sheet = xl_workbook.sheet_by_name(sheet_names[0])

    xl_sheet = xl_workbook.sheet_by_index(0)
    print('Sheet name: %s' % xl_sheet.name)
    row = xl_sheet.row(0)
    #print (row)
    print('(Column #) type:value')
    for idx, cell_obj in enumerate(row):
        cell_type_str = ctype_text.get(cell_obj.ctype, 'unknown type')
        print('(%s) %s %s' % (idx, cell_type_str, cell_obj.value))

    num_cols = xl_sheet.ncols   # Number of columns
    for row_idx in range(0, xl_sheet.nrows):    # Iterate through rows
        print ('-'*40)
        print ('Row: %s' % row_idx)   # Print row number
        for col_idx in range(0, num_cols):  # Iterate through columns
            cell_obj = xl_sheet.cell(row_idx, col_idx)  # Get cell object by row, col
            print ('Column: [%s] cell_obj: [%s]' % (col_idx, cell_obj))
            ws.cell(row=row_idx+1, column=col_idx+1).value =cell_obj.value
    f = 'c:\\OM\\output-' + str(uuid.uuid4())+'.xlsx'
    wbb.save(filename=f)
except:
    print('corupted file')